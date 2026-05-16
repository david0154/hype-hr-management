# salary_monthend.py — Month-End Bulk Salary Generator + Excel Export
# Features:
#   1. Generate salary for ALL active employees in one click
#   2. Live progress bar during generation
#   3. Results table
#   4. Export to Excel (3 sheets: All / Cash / Bank+UPI)
#   5. Save salary slips to Firestore under "salary" collection
#   6. FIX: Auto-reduce advance by amount actually deducted (not full wipe)
# Developed by David | Nexuzy Lab | nexuzylab@gmail.com

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, date
from utils.db import read_all, write, update
from modules.salary import calculate_salary, apply_advance_deduction
import calendar, threading, os

MONTHS = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]


class MonthEndPanel(tk.Frame):
    """Full Month-End salary generation panel."""

    def __init__(self, parent, role="admin"):
        super().__init__(parent, bg="#0d1b2a")
        self.role    = role
        self.results = []
        self._build_ui()

    def _build_ui(self):
        hdr = tk.Frame(self, bg="#1a2740", pady=10)
        hdr.pack(fill="x")
        tk.Label(hdr, text="\U0001f4c5 Month-End Salary Generation",
                 font=("Helvetica",14,"bold"), bg="#1a2740", fg="#f0c040").pack(side="left", padx=12)

        ctrl = tk.Frame(self, bg="#0d1b2a", pady=6)
        ctrl.pack(fill="x", padx=12)

        tk.Label(ctrl, text="Month:", bg="#0d1b2a", fg="#ccc").pack(side="left")
        prev_month = datetime.now().month - 1 or 12
        self.month_var = tk.StringVar(value=MONTHS[prev_month - 1])
        ttk.Combobox(ctrl, textvariable=self.month_var, values=MONTHS,
                     width=12, state="readonly").pack(side="left", padx=4)

        tk.Label(ctrl, text="Year:", bg="#0d1b2a", fg="#ccc").pack(side="left", padx=(10,0))
        prev_year = datetime.now().year if prev_month != 12 else datetime.now().year - 1
        self.year_var = tk.StringVar(value=str(prev_year))
        tk.Entry(ctrl, textvariable=self.year_var, width=6,
                 bg="#1e3a5f", fg="white", insertbackground="white",
                 relief="flat", bd=4).pack(side="left", padx=4)

        tk.Button(ctrl, text="\u26a1 Generate All Salaries",
                  command=self._confirm_generate,
                  bg="#f77f00", fg="white", font=("Arial",10,"bold"),
                  relief="flat", padx=14, pady=5, cursor="hand2").pack(side="left", padx=12)

        tk.Button(ctrl, text="\U0001f4ca Export Excel",
                  command=self._export_excel,
                  bg="#27ae60", fg="white", font=("Arial",10,"bold"),
                  relief="flat", padx=14, pady=5, cursor="hand2").pack(side="left", padx=4)

        tk.Button(ctrl, text="\U0001f504 Reload",
                  command=self._load_existing,
                  bg="#555", fg="white", relief="flat",
                  padx=10, pady=5).pack(side="left", padx=4)

        self.progress_frame = tk.Frame(self, bg="#0d1b2a")
        self.progress_frame.pack(fill="x", padx=12, pady=2)
        self.progress_lbl = tk.Label(self.progress_frame, text="",
                                     bg="#0d1b2a", fg="#aaa", font=("Arial",9))
        self.progress_lbl.pack(side="left")
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.progress_frame, variable=self.progress_var,
                                            maximum=100, length=400, mode="determinate")
        self.progress_bar.pack(side="left", padx=8)
        self.progress_frame.pack_forget()

        self.summary_frm = tk.Frame(self, bg="#132030", pady=8)
        self.summary_frm.pack(fill="x", padx=12, pady=2)
        self.summary_lbl = tk.Label(self.summary_frm, text="  Select month & click Generate",
                                    bg="#132030", fg="#aaa", font=("Arial",9), justify="left")
        self.summary_lbl.pack(anchor="w", padx=8)

        filt = tk.Frame(self, bg="#0d1b2a", pady=3)
        filt.pack(fill="x", padx=12)
        tk.Label(filt, text="Filter:", bg="#0d1b2a", fg="#ccc", font=("Arial",9)).pack(side="left")
        self.filter_var = tk.StringVar(value="All")
        for lbl in ["All", "CASH", "BANK TRANSFER", "UPI", "CHEQUE"]:
            tk.Radiobutton(filt, text=lbl, variable=self.filter_var, value=lbl,
                           bg="#0d1b2a", fg="#ccc", selectcolor="#1a2740",
                           activebackground="#0d1b2a", activeforeground="white",
                           command=self._apply_filter).pack(side="left", padx=4)

        cols = ("no","emp_id","name","dept","mode","present","absent",
                "base","ot_pay","bonus","advance","adv_deducted","remaining_adv","gross","net","status")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        col_cfg = [
            ("no",           "#",             35,  "center"),
            ("emp_id",       "Emp ID",         90,  "center"),
            ("name",         "Name",          150,  "w"),
            ("dept",         "Department",    110,  "w"),
            ("mode",         "Pay Mode",       90,  "center"),
            ("present",      "Present",        65,  "center"),
            ("absent",       "Absent",         60,  "center"),
            ("base",         "Base",           90,  "e"),
            ("ot_pay",       "OT Pay",         80,  "e"),
            ("bonus",        "Bonus",          80,  "e"),
            ("advance",      "Advance",        80,  "e"),
            ("adv_deducted", "Adv.Deducted",   90,  "e"),   # FIX: new column
            ("remaining_adv","Adv.Remaining",  90,  "e"),   # FIX: new column
            ("gross",        "Gross",          90,  "e"),
            ("net",          "Net Pay",         95,  "e"),
            ("status",       "Status",          70,  "center"),
        ]
        for cid, lbl, w, anchor in col_cfg:
            self.tree.heading(cid, text=lbl)
            self.tree.column(cid, width=w, anchor=anchor)

        scroll_y = ttk.Scrollbar(self, orient="vertical",   command=self.tree.yview)
        scroll_x = ttk.Scrollbar(self, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.tree.pack(fill="both", expand=True, padx=12, pady=4)
        scroll_y.pack(side="right", fill="y")
        scroll_x.pack(side="bottom", fill="x")

        self.tree.tag_configure("cash",  background="#0a2010", foreground="#66ff99")
        self.tree.tag_configure("bank",  background="#0a1020", foreground="#66bbff")
        self.tree.tag_configure("upi",   background="#1a0a20", foreground="#cc99ff")
        self.tree.tag_configure("error", background="#200a0a", foreground="#ff6666")

        self._load_existing()

    def _confirm_generate(self):
        month_name = self.month_var.get()
        try:
            year = int(self.year_var.get().strip())
        except ValueError:
            messagebox.showerror("Error", "Enter a valid year."); return
        if not messagebox.askyesno(
                "\u26a1 Confirm Bulk Generation",
                f"Generate salary slips for ALL active employees\n"
                f"for  {month_name} {year} ?\n\n"
                f"\u2022 Advance deducted from gross (partial if gross < advance)\n"
                f"\u2022 Remaining advance carries to next month\n"
                f"\u2022 Existing slips for this month will be overwritten",
                parent=self):
            return
        threading.Thread(target=self._generate_all, daemon=True).start()

    def _generate_all(self):
        month_idx  = MONTHS.index(self.month_var.get()) + 1
        try:
            year = int(self.year_var.get().strip())
        except ValueError:
            return
        month_str  = f"{year}-{month_idx:02d}"

        self.after(0, self._show_progress)

        employees = read_all("employees", "status", "active")
        total     = len(employees)
        if total == 0:
            self.after(0, lambda: messagebox.showinfo("No Employees", "No active employees found."))
            self.after(0, self._hide_progress)
            return

        results = []
        for idx, emp in enumerate(employees):
            pct = int((idx / total) * 100)
            msg = f"Processing {idx+1}/{total}: {emp.get('name','')}"
            self.after(0, lambda p=pct, m=msg: self._update_progress(p, m))
            try:
                from google.cloud.firestore_v1.base_query import FieldFilter
                from utils.firebase_config import get_db
                docs = get_db().collection("sessions") \
                    .where(filter=FieldFilter("employee_id", "==", emp["employee_id"])).stream()
                sessions = [
                    doc.to_dict() for doc in docs
                    if doc.to_dict().get("date", "").startswith(month_str)
                ]
                result   = calculate_salary(emp, sessions, month_idx, year)
                slip_key = f"{emp['employee_id']}_{year}_{month_idx:02d}"
                write("salary", slip_key, {
                    **result,
                    "generated_at": datetime.now().isoformat(),
                    "month_label":  f"{self.month_var.get()} {year}",
                })
                # FIX: reduce advance by exactly what was deducted — not full wipe
                if result.get("advance_deducted", 0) > 0:
                    apply_advance_deduction(
                        employee_id=emp["employee_id"],
                        deducted=result["advance_deducted"],
                        original=result["advance"],
                    )
                result["_status"] = "\u2705 Done"
            except Exception as ex:
                result = {
                    "employee_id":  emp.get("employee_id",""),
                    "name":         emp.get("name",""),
                    "department":   emp.get("department",""),
                    "payment_mode": emp.get("payment_mode","CASH"),
                    "full_days":    0, "absent_days": 0,
                    "base_salary":  0, "ot_pay": 0, "annual_bonus": 0,
                    "advance":      0, "advance_deducted": 0,
                    "remaining_advance": 0, "final_salary": 0,
                    "_status":      f"\u274c {ex}",
                }
            results.append(result)

        self.results = results
        self.after(0, lambda: self._update_progress(100, f"\u2705 Done \u2014 {total} employees processed"))
        self.after(300, self._hide_progress)
        self.after(350, self._populate_table)

    def _show_progress(self):
        self.progress_frame.pack(fill="x", padx=12, pady=4)

    def _hide_progress(self):
        self.progress_frame.pack_forget()

    def _update_progress(self, pct, msg):
        self.progress_var.set(pct)
        self.progress_lbl.config(text=msg)

    def _load_existing(self):
        month_idx = MONTHS.index(self.month_var.get()) + 1
        try:
            year = int(self.year_var.get().strip())
        except ValueError:
            return
        all_slips = read_all("salary")
        self.results = [
            s for s in all_slips
            if str(s.get("year","")) == str(year)
            and str(s.get("month","")) == str(month_idx)
        ]
        for r in self.results:
            r.setdefault("_status", "\u2705 Saved")
        self._populate_table()

    def _populate_table(self, mode_filter="All"):
        mode_filter = self.filter_var.get()
        self.tree.delete(*self.tree.get_children())
        total_gross = total_net = total_cash = total_bank = 0.0
        cash_count  = bank_count = 0
        shown       = 0

        for idx, r in enumerate(self.results, start=1):
            mode = (r.get("payment_mode") or "CASH").upper()
            if mode_filter != "All" and mode != mode_filter:
                continue
            shown += 1
            full    = r.get("full_days", 0)
            half    = r.get("half_days", 0)
            present = full + half * 0.5
            absent  = r.get("absent_days", 0)
            base    = r.get("base_salary", 0)
            ot_pay  = r.get("ot_pay", 0)
            bonus   = r.get("annual_bonus", 0)
            advance = r.get("advance", 0)
            adv_ded = r.get("advance_deducted", 0)       # FIX
            adv_rem = r.get("remaining_advance", 0)      # FIX
            gross   = r.get("attendance_salary", 0) + ot_pay + bonus
            net     = r.get("final_salary", 0)
            status  = r.get("_status", "")

            total_gross += gross
            total_net   += net
            if mode == "CASH":
                total_cash += net; cash_count += 1
                tag = "cash"
            elif mode in ("BANK TRANSFER", "UPI", "CHEQUE"):
                total_bank += net; bank_count += 1
                tag = "bank" if mode == "BANK TRANSFER" else "upi"
            else:
                tag = "cash"
            if "\u274c" in status: tag = "error"

            self.tree.insert("", "end", tags=(tag,), values=(
                shown,
                r.get("employee_id",""),
                r.get("name",""),
                r.get("department",""),
                mode,
                f"{present:.1f}",
                f"{absent:.1f}",
                f"\u20b9{base:,.0f}",
                f"\u20b9{ot_pay:,.0f}",
                f"\u20b9{bonus:,.0f}",
                f"\u20b9{advance:,.0f}",
                f"\u20b9{adv_ded:,.0f}",   # FIX
                f"\u20b9{adv_rem:,.0f}",   # FIX
                f"\u20b9{gross:,.0f}",
                f"\u20b9{net:,.0f}",
                status,
            ))

        self.summary_lbl.config(
            text=f"  \U0001f4ca Total Employees: {shown}   "
                 f"\U0001f4b0 Total Gross: \u20b9{total_gross:,.0f}   "
                 f"\u2705 Total Net: \u20b9{total_net:,.0f}   "
                 f"\u2502   \U0001f4b5 Cash: {cash_count} emp  \u20b9{total_cash:,.0f}   "
                 f"\U0001f3e6 Bank/UPI: {bank_count} emp  \u20b9{total_bank:,.0f}"
        )

    def _apply_filter(self):
        self._populate_table(self.filter_var.get())

    def _export_excel(self):
        if not self.results:
            messagebox.showwarning("No Data",
                "No salary data to export.\nPlease generate salaries first.",
                parent=self)
            return

        month_name = self.month_var.get()
        try:
            year = int(self.year_var.get().strip())
        except ValueError:
            return

        default_name = f"Salary_{month_name}_{year}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Save Excel Report",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import (Font, PatternFill, Alignment,
                                         Border, Side)
            from openpyxl.utils import get_column_letter
        except ImportError:
            messagebox.showerror("Missing Library",
                "openpyxl not installed.\n\nRun:  pip install openpyxl",
                parent=self)
            return

        wb = openpyxl.Workbook()

        HDR_FILL   = PatternFill("solid", fgColor="1A2740")
        TITLE_FILL = PatternFill("solid", fgColor="F77F00")
        CASH_FILL  = PatternFill("solid", fgColor="E8F5E9")
        BANK_FILL  = PatternFill("solid", fgColor="E3F2FD")
        UPI_FILL   = PatternFill("solid", fgColor="F3E5F5")
        TOTAL_FILL = PatternFill("solid", fgColor="FFF9C4")
        HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
        TITLE_FONT = Font(bold=True, color="FFFFFF", size=12)
        TOTAL_FONT = Font(bold=True, color="000000", size=10)
        thin       = Side(style="thin")
        BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)
        CENTER     = Alignment(horizontal="center", vertical="center")
        RIGHT      = Alignment(horizontal="right",  vertical="center")
        INR        = '\u20b9#,##0.00'

        def style_header_row(ws, row, cols_count):
            for col in range(1, cols_count+1):
                cell = ws.cell(row=row, column=col)
                cell.fill = HDR_FILL; cell.font = HDR_FONT
                cell.border = BORDER; cell.alignment = CENTER

        def style_title(ws, row, col, text, span):
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col+span-1)
            c = ws.cell(row=row, column=col, value=text)
            c.fill = TITLE_FILL; c.font = TITLE_FONT; c.alignment = CENTER

        def set_col_widths(ws, widths):
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

        def money_cell(ws, row, col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.number_format = INR; c.alignment = RIGHT; c.border = BORDER
            return c

        def text_cell(ws, row, col, val, align=CENTER):
            c = ws.cell(row=row, column=col, value=val)
            c.alignment = align; c.border = BORDER
            return c

        # Sheet 1 — All Employees
        ws1 = wb.active
        ws1.title = "All Employees"
        ws1.row_dimensions[1].height = 24
        style_title(ws1, 1, 1, f"HYPE HR MANAGEMENT \u2014 Salary Report  {month_name} {year}", 16)
        ws1.cell(row=2, column=1,
                 value=f"Generated: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}")
        ws1.cell(row=2, column=1).font = Font(italic=True, size=9, color="555555")

        headers1 = ["#","Emp ID","Name","Department","Pay Mode",
                    "Present","Absent","Base Salary","Att. Salary",
                    "OT Pay","Bonus","Advance","Adv.Deducted","Adv.Remaining",  # FIX: 2 new cols
                    "Gross Pay","Net Pay","Status"]
        for col, h in enumerate(headers1, start=1):
            ws1.cell(row=3, column=col, value=h)
        style_header_row(ws1, 3, len(headers1))

        grand_gross = grand_net = 0.0
        for i, r in enumerate(self.results, start=1):
            row = i + 3
            mode    = (r.get("payment_mode") or "CASH").upper()
            present = r.get("full_days",0) + r.get("half_days",0)*0.5
            absent  = r.get("absent_days",0)
            base    = float(r.get("base_salary",0))
            att_sal = float(r.get("attendance_salary",0))
            ot_pay  = float(r.get("ot_pay",0))
            bonus   = float(r.get("annual_bonus",0))
            advance = float(r.get("advance",0))
            adv_ded = float(r.get("advance_deducted",0))
            adv_rem = float(r.get("remaining_advance",0))
            gross   = att_sal + ot_pay + bonus
            net     = float(r.get("final_salary",0))
            grand_gross += gross; grand_net += net

            fill = CASH_FILL if mode=="CASH" else (BANK_FILL if mode=="BANK TRANSFER" else UPI_FILL)
            LEFT = Alignment(horizontal="left", vertical="center")

            text_cell(ws1, row, 1, i)
            text_cell(ws1, row, 2, r.get("employee_id",""))
            text_cell(ws1, row, 3, r.get("name",""),       align=LEFT)
            text_cell(ws1, row, 4, r.get("department",""), align=LEFT)
            text_cell(ws1, row, 5, mode)
            text_cell(ws1, row, 6, round(present,1))
            text_cell(ws1, row, 7, round(absent,1))
            money_cell(ws1, row, 8, base)
            money_cell(ws1, row, 9, att_sal)
            money_cell(ws1, row, 10, ot_pay)
            money_cell(ws1, row, 11, bonus)
            money_cell(ws1, row, 12, advance)
            money_cell(ws1, row, 13, adv_ded)    # FIX
            money_cell(ws1, row, 14, adv_rem)    # FIX
            money_cell(ws1, row, 15, gross)
            money_cell(ws1, row, 16, net)
            text_cell(ws1, row, 17, r.get("_status",""))
            for col in range(1, 18): ws1.cell(row=row, column=col).fill = fill

        tr = len(self.results) + 4
        ws1.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=14)
        c = ws1.cell(row=tr, column=1, value="GRAND TOTAL")
        c.font = TOTAL_FONT; c.fill = TOTAL_FILL; c.alignment = RIGHT; c.border = BORDER
        money_cell(ws1, tr, 15, grand_gross).font = TOTAL_FONT
        money_cell(ws1, tr, 16, grand_net).font   = TOTAL_FONT
        for col in range(1, 18): ws1.cell(row=tr, column=col).fill = TOTAL_FILL
        set_col_widths(ws1, [4,11,22,16,14,8,8,13,13,11,11,11,12,12,13,13,10])

        # Sheet 2 — Cash
        ws2 = wb.create_sheet("Cash Payments")
        style_title(ws2, 1, 1, f"\U0001f4b5 CASH PAYMENT LIST \u2014 {month_name} {year}", 9)
        headers2 = ["#","Emp ID","Name","Department","Designation",
                    "Net Pay (\u20b9)","Signature","Date","Remarks"]
        for col, h in enumerate(headers2, start=1):
            ws2.cell(row=2, column=col, value=h)
        style_header_row(ws2, 2, len(headers2))

        cash_emps = sorted(
            [r for r in self.results if (r.get("payment_mode") or "CASH").upper() == "CASH"],
            key=lambda x: float(x.get("final_salary",0)), reverse=True
        )
        cash_total = 0.0
        LEFT = Alignment(horizontal="left", vertical="center")
        for i, r in enumerate(cash_emps, start=1):
            row = i + 2
            net = float(r.get("final_salary",0))
            cash_total += net
            text_cell(ws2, row, 1, i)
            text_cell(ws2, row, 2, r.get("employee_id",""))
            text_cell(ws2, row, 3, r.get("name",""),        align=LEFT)
            text_cell(ws2, row, 4, r.get("department",""),  align=LEFT)
            text_cell(ws2, row, 5, r.get("designation",""), align=LEFT)
            money_cell(ws2, row, 6, net)
            text_cell(ws2, row, 7, ""); text_cell(ws2, row, 8, ""); text_cell(ws2, row, 9, "")
            for col in range(1, 10): ws2.cell(row=row, column=col).fill = CASH_FILL

        tr2 = len(cash_emps) + 3
        ws2.merge_cells(start_row=tr2, start_column=1, end_row=tr2, end_column=5)
        c = ws2.cell(row=tr2, column=1,
                     value=f"TOTAL CASH TO ARRANGE: \u20b9{cash_total:,.2f}")
        c.font = Font(bold=True, size=12, color="CC0000")
        c.fill = TOTAL_FILL; c.alignment = RIGHT; c.border = BORDER
        money_cell(ws2, tr2, 6, cash_total).font = Font(bold=True, size=12, color="CC0000")
        ws2.cell(row=tr2, column=6).fill = TOTAL_FILL
        set_col_widths(ws2, [4,11,22,16,18,14,20,14,16])

        # Sheet 3 — Bank/UPI
        ws3 = wb.create_sheet("Bank & UPI")
        style_title(ws3, 1, 1, f"\U0001f3e6 BANK / UPI TRANSFER LIST \u2014 {month_name} {year}", 8)
        headers3 = ["#","Emp ID","Name","Pay Mode","Net Pay (\u20b9)",
                    "Bank / UPI ID","Account No.","IFSC","Status"]
        for col, h in enumerate(headers3, start=1):
            ws3.cell(row=2, column=col, value=h)
        style_header_row(ws3, 2, len(headers3))

        bank_emps  = [r for r in self.results
                      if (r.get("payment_mode") or "CASH").upper() != "CASH"]
        bank_total = 0.0
        for i, r in enumerate(bank_emps, start=1):
            row  = i + 2
            net  = float(r.get("final_salary",0))
            mode = (r.get("payment_mode") or "").upper()
            bank_total += net
            fill = BANK_FILL if mode=="BANK TRANSFER" else UPI_FILL
            text_cell(ws3, row, 1, i)
            text_cell(ws3, row, 2, r.get("employee_id",""))
            text_cell(ws3, row, 3, r.get("name",""), align=Alignment(horizontal="left",vertical="center"))
            text_cell(ws3, row, 4, mode)
            money_cell(ws3, row, 5, net)
            text_cell(ws3, row, 6, r.get("upi_id","") or r.get("bank_upi",""))
            text_cell(ws3, row, 7, r.get("account_number",""))
            text_cell(ws3, row, 8, r.get("ifsc",""))
            text_cell(ws3, row, 9, "Pending")
            for col in range(1, 10): ws3.cell(row=row, column=col).fill = fill

        tr3 = len(bank_emps) + 3
        ws3.merge_cells(start_row=tr3, start_column=1, end_row=tr3, end_column=4)
        c = ws3.cell(row=tr3, column=1, value="TOTAL BANK/UPI TRANSFER")
        c.font = TOTAL_FONT; c.fill = TOTAL_FILL; c.alignment = RIGHT; c.border = BORDER
        money_cell(ws3, tr3, 5, bank_total).font = TOTAL_FONT
        ws3.cell(row=tr3, column=5).fill = TOTAL_FILL
        set_col_widths(ws3, [4,11,22,14,14,20,18,12,10])

        wb.save(path)
        messagebox.showinfo(
            "\u2705 Export Complete",
            f"Excel saved:\n{path}\n\n"
            f"\U0001f4c4 Sheet 1 \u2014 All Employees ({len(self.results)} records)\n"
            f"\U0001f4b5 Sheet 2 \u2014 Cash Payments ({len(cash_emps)} employees  \u20b9{cash_total:,.0f})\n"
            f"\U0001f3e6 Sheet 3 \u2014 Bank/UPI ({len(bank_emps)} employees  \u20b9{bank_total:,.0f})",
            parent=self
        )
        try:
            os.startfile(path)
        except Exception:
            try:
                import subprocess; subprocess.Popen(["xdg-open", path])
            except Exception:
                pass
